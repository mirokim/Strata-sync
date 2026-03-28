#!/usr/bin/env python3
"""
xlsx_to_md.py — §4.4 XLSX → Obsidian Markdown conversion

- 1 sheet = 1 ## {sheet_name} section
- Cell data → Markdown table (merged cells: use first cell value)
- 빈 시트·숨겨진 시트 Skipped
- 열이 너무 많으면(20열+) 첫 20열만 추출 + 경고
- 행이 1000개 이상인 시트는 허브 노트(요약 + 행 수 기록)로 처리

Usage:
  python xlsx_to_md.py <src_dir_or_file> <active_dir> <attachments_dir>
"""

import re
import sys
import argparse
import warnings
from pathlib import Path
from datetime import datetime

warnings.filterwarnings('ignore')

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl --break-system-packages")
    sys.exit(1)


MAX_COLS = 20      # 이 열 수 초과 시 잘라냄
MAX_ROWS = 1000    # 이 행 수 초과 시 허브 요약 처리
MAX_CELL_LEN = 200 # 셀 내용 최대 길이


def cell_str(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    # Escape pipes, remove newlines
    s = s.replace('|', '\\|').replace('\n', ' ').replace('\r', '')
    return s[:MAX_CELL_LEN] if len(s) > MAX_CELL_LEN else s


def sheet_to_md_table(ws) -> tuple[str, int]:
    """Sheet → Markdown table string. Returns (md_text, row_count)."""
    # Collect valid rows (excluding completely empty rows)
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            all_rows.append(row)

    if not all_rows:
        return '', 0

    total_rows = len(all_rows)
    n_cols = min(max(len(r) for r in all_rows), MAX_COLS)

    # 1000행 초과: 허브 요약만
    if total_rows > MAX_ROWS:
        # 헤더 + 처음 5행 + 마지막 3행 샘플만
        header = all_rows[0][:n_cols]
        sample_rows = all_rows[1:6] + all_rows[-3:]
        note = f"\n> ⚠️ 총 {total_rows}행 (1,000행 초과). 아래는 샘플 {len(sample_rows)}행.\n"
    else:
        header = all_rows[0][:n_cols]
        sample_rows = all_rows[1:]
        note = ''

    lines = []
    if note:
        lines.append(note)

    # 헤더
    h_cells = [cell_str(c) or f'Col{i+1}' for i, c in enumerate(header)]
    lines.append('| ' + ' | '.join(h_cells) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(h_cells)) + ' |')

    # 데이터 행
    for row in sample_rows:
        cells = [cell_str(row[i] if i < len(row) else None) for i in range(n_cols)]
        if any(cells):
            lines.append('| ' + ' | '.join(cells) + ' |')

    col_note = f"\n> 열 수: {max(len(r) for r in all_rows)} → 앞 {n_cols}열만 표시\n" \
               if max(len(r) for r in all_rows) > MAX_COLS else ''

    return '\n'.join(lines) + col_note, total_rows


def xlsx_to_md(xlsx_path: Path, active_dir: Path, attachments_dir: Path) -> bool:
    """Convert a single XLSX to MD. Returns True on success."""
    stem = xlsx_path.stem
    safe_stem = re.sub(r'[<>:"/\\|?*]', '_', stem)
    out_md = active_dir / f"{safe_stem}.md"

    if out_md.exists():
        return False  # 이미 변환됨

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as e:
        print(f"  ✗ {xlsx_path.name}: {e}")
        return False

    sections = []
    total_sheets = 0
    total_rows = 0

    for ws in wb.worksheets:
        # Skip hidden sheets
        if ws.sheet_state == 'hidden':
            continue
        table_md, n_rows = sheet_to_md_table(ws)
        if not table_md:
            continue
        total_sheets += 1
        total_rows += n_rows
        sections.append(f"## {ws.title}\n\n{table_md}")

    wb.close()

    if not sections:
        return False  # 내용 없음

    # frontmatter
    try:
        mtime = xlsx_path.stat().st_mtime
        date_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d')
    except Exception:
        date_str = datetime.now().strftime('%Y-%m-%d')

    m_date = re.search(r'(\d{4})[_\-](\d{2})[_\-](\d{2})', stem)
    if m_date:
        date_str = f"{m_date.group(1)}-{m_date.group(2)}-{m_date.group(3)}"
    else:
        m_date2 = re.search(r'(\d{4})(\d{2})(\d{2})', stem)
        if m_date2:
            date_str = f"{m_date2.group(1)}-{m_date2.group(2)}-{m_date2.group(3)}"

    doc_type = 'spec'
    if any(kw in stem for kw in ['일정', '관리', '계획']):
        doc_type = 'spec'
    elif any(kw in stem for kw in ['레퍼런스', '분석']):
        doc_type = 'reference'

    body = '\n\n---\n\n'.join(sections)

    md_content = f"""---
title: "{stem}"
date: {date_str}
type: {doc_type}
status: active
tags: []
source: "{xlsx_path.name}"
origin: xlsx
sheets: {total_sheets}
total_rows: {total_rows}
---

# {stem}

> 원본: `{xlsx_path.name}` ({total_sheets}개 시트, 총 {total_rows}행)

{body}
"""

    out_md.write_text(md_content, encoding='utf-8')
    return True


def process_directory(src_dir: Path, active_dir: Path, attachments_dir: Path) -> tuple:
    xlsxs = list(src_dir.rglob('*.xlsx'))
    # Remove duplicate filenames (by stem)
    seen_stems = set()
    unique_xlsxs = []
    for x in xlsxs:
        safe = re.sub(r'[<>:"/\\|?*]', '_', x.stem)
        if safe not in seen_stems:
            seen_stems.add(safe)
            unique_xlsxs.append(x)

    success = skip = fail = 0
    for xlsx in unique_xlsxs:
        safe = re.sub(r'[<>:"/\\|?*]', '_', xlsx.stem)
        if (active_dir / f"{safe}.md").exists():
            skip += 1
            continue
        ok = xlsx_to_md(xlsx, active_dir, attachments_dir)
        if ok:
            success += 1
        else:
            fail += 1

    return success, fail, skip


def main():
    parser = argparse.ArgumentParser(description='§4.4 XLSX → Markdown 변환')
    parser.add_argument('src', help='XLSX file or directory path')
    parser.add_argument('active_dir', help='active/ folder path')
    parser.add_argument('attachments_dir', help='attachments/ folder path')
    args = parser.parse_args()

    src = Path(args.src)
    active_dir = Path(args.active_dir)
    attachments_dir = Path(args.attachments_dir)
    active_dir.mkdir(parents=True, exist_ok=True)

    if src.is_file():
        ok = xlsx_to_md(src, active_dir, attachments_dir)
        print('✅ 변환 Complete' if ok else '⚠️ Skipped')
    elif src.is_dir():
        total = len(list(src.rglob('*.xlsx')))
        print(f"XLSX Starting conversion: 총 {total}개 발견")
        success, fail, skip = process_directory(src, active_dir, attachments_dir)
        print(f"\n=== §4.4 XLSX 변환 Complete ===")
        print(f"  Success: {success}개")
        print(f"  Fail: {fail}개")
        print(f"  스킵: {skip}개 (중복 파일명 포함)")
    else:
        print(f"Error: {src} not found")
        sys.exit(1)


if __name__ == '__main__':
    main()
